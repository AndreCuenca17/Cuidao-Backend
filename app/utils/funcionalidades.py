from collections import defaultdict
from geopy.distance import geodesic
import math
import geopandas as gpd
import networkx as nx
import json
import folium
import openpyxl
import pandas as pd
import requests
from shapely import LineString

geolocalizacion = {"latitud": None, "longitud": None}
ubicacion_seleccionada = {"latitud": None, "longitud": None}

def obtener_ubicacion():
    """
    Devuelve las coordenadas almacenadas globalmente.
    Si no están disponibles, devuelve un error.
    """
    global geolocalizacion
    if geolocalizacion["latitud"] is None or geolocalizacion["longitud"] is None:
        print("No se han obtenido ubicacion")
        return {"error": "No se han obtenido ubicacion"}
    
    print("Ubicacion obtenida")
    return {
        "latitud": geolocalizacion["latitud"],
        "longitud": geolocalizacion["longitud"]
    }


def crear_mapa_vacio():
    ubicacion = obtener_ubicacion()

    mapaCuidao = folium.Map(
        location=[ubicacion["latitud"], ubicacion["longitud"]], zoom_start=12
    )
    return mapaCuidao


def obtener_ultima_fila(sheet):
    # Encuentra la última fila con datos en la primera columna
    for row in range(
        2, sheet.max_row + 1
    ):  # Comienza desde la fila 2, asumiendo que la primera tiene encabezados
        if sheet.cell(row=row, column=1).value is None:
            return row
    return sheet.max_row + 1


def cargar_datos_y_contar_distritos(archivo_excel="data/reportes_delitos.xlsx"):
    # Carga el archivo Excel y selecciona la hoja activa
    workbook = openpyxl.load_workbook(archivo_excel, read_only=True)
    sheet = workbook.active

    # Inicializa contenedores
    conteo_distritos = defaultdict(int)
    datos = []

    # Itera sobre las filas relevantes, evitando el encabezado
    for fila in sheet.iter_rows(min_row=2, values_only=True):
        distrito = fila[5]  # Índice 5 corresponde a la columna de distrito
        # Incrementa el conteo del distrito
        conteo_distritos[distrito] += 1

        # Agrega los datos formateados
        datos.append({
            "dni": fila[0],
            "latitud": fila[1],
            "longitud": fila[2],
            "departamento": fila[3],
            "provincia": fila[4],
            "distrito": distrito,
            "tipo_robo": fila[6],
            "descripcion": fila[7],
            "fecha": fila[8],
            "hora": fila[9],
        })

    return datos, conteo_distritos


def obtener_geojson_distritos(archivo_geojson="data/distritos.geojson"):
    with open(archivo_geojson, "r") as f:
        dataGeoJson = json.load(f)
    return dataGeoJson


def cargar_calles_json(file="data/callePrincipal.geojson"):
    # Cargar archivo geojson
    gdf = gpd.read_file(file)
    G = nx.Graph()

    # Iterar sobre las filas del GeoDataFrame
    for _, row in gdf.iterrows():
        # Procesar solo las geometrías de tipo LineString
        if isinstance(row.geometry, LineString):
            # Extraer y convertir coordenadas (invertir latitud y longitud)
            coords = [(lon, lat) for lat, lon in row.geometry.coords]
            # Agregar aristas al grafo entre nodos consecutivos
            for i in range(len(coords) - 1):
                node1 = coords[i]
                node2 = coords[i + 1]
                # Agregar borde y nombre de calle si existe
                G.add_edge(node1, node2, street_name=row.get("name", "Unknown"))

    return G


def euclidean_distance(point1, point2):
    return math.sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)


def find_closest_node(graph, point):
    closest_node = None
    min_distance = float("inf")  # Inicializar con un valor muy grande

    for node in graph.nodes:
        dist = euclidean_distance(node, point)
        if dist < min_distance:
            min_distance = dist
            closest_node = node

    return closest_node


def encontrar_comisaria_mas_cercana(mi_ubicacion, comisarias):

    # Obtener la ubicación del cliente
    # mi_ubicacion = (ubicacion["latitud"], ubicacion["longitud"])  #Descomentar cuando se pueda ejecutar en un servidor local
    mi_ubicacion = mi_ubicacion

    # Calcular distancias entre la ubicación del cliente y las comisarias_diccionario en nuestro diccionario
    distancias = {}
    for comisaria in comisarias:
        nombre = comisaria["name"]
        coords = (comisaria["lat"], comisaria["lng"])
        distancias[nombre] = geodesic(mi_ubicacion, coords).km

    # Determinar la comisaría más cercana
    comisaria_mas_cercana = min(distancias, key=distancias.get)

    return comisaria_mas_cercana


def ConvertirDictToList(
    dicti: dict, lista: list
):  # Para que se ponga en el mapa, se necesita correr una Lista
    for nombre, coordenadas in dicti.items():
        latitud, longitud = coordenadas[0].split(",")
        latitud = float(latitud)
        longitud = float(longitud)
        lista.append({"name": nombre, "lat": latitud, "lng": longitud})


def CargarDataToDict(file_path: str, data_dict: dict, columna1: str, columna2: str):
    datos_excel = {}
    dataframe = pd.read_excel(file_path)
    for _, row in dataframe.iterrows():
        if row[columna1] not in datos_excel:
            datos_excel[row[columna1]] = []
        datos_excel[row[columna1]].append(row[columna2])
    data_dict.update(datos_excel)
